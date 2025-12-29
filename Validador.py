"""
Script de Validação de Colaboradores - Sistema Cebraspe
Valida dados de colaboradores via checkbox
"""

import os
import time
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


def _js_find_visible_by_text(driver, text: str, partial: bool = True, selectors: str = 'a,button,span,li,div'):
    """Retorna o primeiro elemento visível cujo texto corresponde (parcial/exato)."""
    js = r'''
    const txt = arguments[0].toLowerCase();
    const partial = arguments[1];
    const selectors = arguments[2];
    const nodes = document.querySelectorAll(selectors);
    function isVisible(el){
        const cs = window.getComputedStyle(el);
        if(!el.offsetParent && cs.position !== 'fixed') return false;
        return cs.visibility !== 'hidden' && cs.display !== 'none' && el.offsetWidth > 0 && el.offsetHeight > 0;
    }
    for(const el of nodes){
        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
        if(!t) continue;
        if(partial ? t.includes(txt) : t === txt){
            if(isVisible(el)) return el;
        }
    }
    return null;
    '''
    return driver.execute_script(js, text, partial, selectors)


# ========== CONFIGURAÇÕES ==========
LOGIN_URL = "https://extranet.cebraspe.org.br/Sinapse/aspx/autenticarusuario/AutenticarUsuario.aspx"
PLANILHA_PATH = r"C:\Users\victor.vasconcelos\Downloads\Demanda_divisão lanche PPL.xlsx"


def validar_colaborador(driver, cpf):
    """Valida um colaborador específico"""
    try:
        wait = WebDriverWait(driver, 15)
        time.sleep(2)
        
        # Preenche campo CPF
        print(f"[VALIDAÇÃO] Preenchendo CPF: {cpf}")
        cpf_field = None
        
        try:
            cpf_field = driver.find_element(By.XPATH, '//input[@type="text" and (contains(@name,"CPF") or contains(@id,"CPF"))]')
            print("[VALIDAÇÃO] Campo CPF encontrado!")
        except:
            try:
                cpf_field = driver.find_element(By.XPATH, '(//input[@type="text"])[1]')
                print("[VALIDAÇÃO] Campo CPF encontrado (primeiro input)!")
            except Exception as e:
                print(f"[ERRO] Campo CPF não encontrado: {e}")
                raise Exception("Campo CPF não encontrado")
        
        cpf_field.clear()
        time.sleep(0.5)
        cpf_field.send_keys(cpf)
        print(f"[VALIDAÇÃO] CPF {cpf} digitado.")
        time.sleep(1)
        
        # Marca checkbox Visualizar todas as Cidades
        print("[VALIDAÇÃO] Marcando checkbox 'Visualizar todas as Cidades'...")
        checkbox_cidades = None
        try:
            checkbox_cidades = driver.find_element(
                By.XPATH, '//input[@type="checkbox" and (contains(translate(following-sibling::text()[1],"VISUALIZARTODASCIDADES","visualizartodascidades"),"visualizar") or contains(translate(../text(),"VISUALIZARTODASCIDADES","visualizartodascidades"),"visualizar"))]'
            )
            print("[VALIDAÇÃO] Checkbox encontrado por texto!")
        except:
            try:
                checkboxes = driver.find_elements(By.XPATH, '//input[@type="checkbox"]')
                if len(checkboxes) > 0:
                    checkbox_cidades = checkboxes[0]
                    print("[VALIDAÇÃO] Checkbox encontrado (primeiro da página)!")
            except:
                pass
        
        if checkbox_cidades and not checkbox_cidades.is_selected():
            try:
                checkbox_cidades.click()
                print("[VALIDAÇÃO] Checkbox marcado.")
            except:
                driver.execute_script("arguments[0].click();", checkbox_cidades)
                print("[VALIDAÇÃO] Checkbox marcado (JS).")
        time.sleep(1)
        
        # Clica em Pesquisar
        print("[VALIDAÇÃO] Clicando em Pesquisar...")
        pesquisar_button = None
        try:
            pesquisar_button = driver.find_element(
                By.XPATH, '//button[contains(text(),"Pesquisar")] | //input[@type="submit" and contains(@value,"Pesquisar")] | //input[@type="button" and contains(@value,"Pesquisar")]'
            )
            print("[VALIDAÇÃO] Botão Pesquisar encontrado!")
        except:
            pesquisar_button = _js_find_visible_by_text(driver, 'pesquisar', partial=True, selectors='button,input[type=submit],input[type=button]')
            if pesquisar_button:
                print("[VALIDAÇÃO] Botão Pesquisar encontrado (JS)!")
        
        if not pesquisar_button:
            print("[ERRO] Botão Pesquisar não encontrado.")
            raise Exception("Botão Pesquisar não encontrado")
        
        try:
            pesquisar_button.click()
            print("[VALIDAÇÃO] Botão Pesquisar clicado.")
        except:
            driver.execute_script("arguments[0].click();", pesquisar_button)
            print("[VALIDAÇÃO] Botão Pesquisar clicado (JS).")
        
        print("[VALIDAÇÃO] Aguardando resultado da pesquisa...")
        time.sleep(6)
        
        # Marca checkbox Validade na Receita Federal?
        print("[VALIDAÇÃO] Marcando checkbox 'Validade na Receita Federal?'...")
        checkbox_receita = None
        try:
            checkboxes = driver.find_elements(By.XPATH, '//input[@type="checkbox"]')
            print(f"[DEBUG] Total de checkboxes: {len(checkboxes)}")
            
            for idx, cb in enumerate(checkboxes):
                try:
                    parent_text = cb.find_element(By.XPATH, './..').text
                    if 'receita' in parent_text.lower() or 'federal' in parent_text.lower():
                        checkbox_receita = cb
                        print(f"[VALIDAÇÃO] Checkbox 'Receita Federal' encontrado (índice {idx})!")
                        break
                except:
                    continue
            
            if not checkbox_receita and len(checkboxes) > 1:
                checkbox_receita = checkboxes[1]
                print("[VALIDAÇÃO] Usando segundo checkbox.")
        except Exception as e:
            print(f"[ERRO] Erro ao localizar checkbox Receita Federal: {e}")
        
        if checkbox_receita and not checkbox_receita.is_selected():
            try:
                checkbox_receita.click()
                print("[VALIDAÇÃO] Checkbox 'Receita Federal' marcado.")
            except:
                driver.execute_script("arguments[0].click();", checkbox_receita)
                print("[VALIDAÇÃO] Checkbox 'Receita Federal' marcado (JS).")
        time.sleep(1)
        
        # Clica em Gravar
        print("[VALIDAÇÃO] Clicando em Gravar...")
        gravar_button = None
        try:
            gravar_button = driver.find_element(
                By.XPATH, '//button[contains(translate(.,"GRAVAR","gravar"),"gravar")] | //input[@type="submit" and contains(translate(@value,"GRAVAR","gravar"),"gravar")] | //input[@type="button" and contains(translate(@value,"GRAVAR","gravar"),"gravar")]'
            )
        except:
            gravar_button = _js_find_visible_by_text(driver, 'gravar', partial=True, selectors='button,input[type=submit],input[type=button]')
        
        if not gravar_button:
            raise Exception("Botão Gravar não encontrado")
        
        try:
            gravar_button.click()
        except:
            driver.execute_script("arguments[0].click();", gravar_button)
        
        print("[VALIDAÇÃO] Aguardando confirmação...")
        time.sleep(2)
        
        # Trata popup
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            alert_text = ""
            try:
                alert_text = alert.text
            except:
                pass
            print(f"[VALIDAÇÃO] Alert: {alert_text[:50]}...")
            alert.accept()
            print("[VALIDAÇÃO] Alert aceito.")
            time.sleep(1)
        except:
            try:
                ok_button = driver.find_element(
                    By.XPATH, '//button[contains(text(),"OK") or contains(text(),"Ok")] | //input[@type="button" and contains(@value,"OK")]'
                )
                try:
                    ok_button.click()
                except:
                    driver.execute_script("arguments[0].click();", ok_button)
                print("[VALIDAÇÃO] Botão OK clicado.")
                time.sleep(1)
            except:
                print("[VALIDAÇÃO] Nenhum popup encontrado.")
        
        print(f"[VALIDAÇÃO] ✓ CPF {cpf} validado com sucesso!")
        
        # Clica em Voltar
        print("[VALIDAÇÃO] Procurando botão Voltar...")
        voltar_encontrado = False
        
        try:
            time.sleep(1)
            voltar_button = driver.find_element(
                By.XPATH, '//button[normalize-space(text())="Voltar"] | //input[@type="button" and normalize-space(@value)="Voltar"] | //input[@type="submit" and normalize-space(@value)="Voltar"]'
            )
            voltar_encontrado = True
            print("[VALIDAÇÃO] Botão Voltar encontrado!")
        except:
            try:
                voltar_button = driver.find_element(
                    By.XPATH, '//button[contains(text(),"Voltar")] | //input[contains(@value,"Voltar")]'
                )
                voltar_encontrado = True
                print("[VALIDAÇÃO] Botão Voltar encontrado (contains)!")
            except:
                try:
                    voltar_button = _js_find_visible_by_text(driver, 'voltar', partial=False, selectors='button,input[type=button],input[type=submit]')
                    if voltar_button:
                        voltar_encontrado = True
                        print("[VALIDAÇÃO] Botão Voltar encontrado (JS)!")
                except:
                    pass
        
        if voltar_encontrado:
            try:
                print("[VALIDAÇÃO] Clicando em Voltar...")
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", voltar_button)
                time.sleep(0.5)
                voltar_button.click()
                print("[VALIDAÇÃO] Clique executado!")
            except:
                driver.execute_script("arguments[0].click();", voltar_button)
                print("[VALIDAÇÃO] Clique JS executado!")
            
            print("[VALIDAÇÃO] Aguardando tela de pesquisa...")
            time.sleep(2)
            print("[VALIDAÇÃO] Pronto para próxima consulta.")
        else:
            print("[ERRO] Botão Voltar não encontrado!")
        
        return True
        
    except Exception as e:
        print(f"[ERRO] Falha ao validar CPF {cpf}: {e}")
        return False


def ler_cpfs_planilha(caminho_planilha):
    """Lê CPFs da planilha Excel (coluna B) e retorna lista com linha e CPF"""
    try:
        if not os.path.exists(caminho_planilha):
            print(f"[ERRO] Planilha não encontrada: {caminho_planilha}")
            return []

        print(f"[PLANILHA] Carregando planilha: {caminho_planilha}")
        wb = load_workbook(caminho_planilha, data_only=True, read_only=True)
        ws = wb.active
        
        cpfs_para_processar = []
        
        # Itera valores por linha (colunas B=índice 1, F=índice 5)
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            cpf = row[1]
            status = row[5] if len(row) > 5 else None
            
            if cpf and str(status).strip().upper() != "FEITO":
                cpf_limpo = str(cpf).strip().replace('.', '').replace('-', '').replace(' ', '')
                if cpf_limpo:
                    cpfs_para_processar.append({
                        'linha': idx,
                        'cpf': cpf_limpo
                    })
        
        wb.close()
        print(f"[PLANILHA] Encontrados {len(cpfs_para_processar)} CPFs para processar.")
        return cpfs_para_processar
        
    except Exception as e:
        print(f"[ERRO] Falha ao ler planilha: {e}")
        return []


def atualizar_status_planilha_batch(wb, caminho_planilha, contador, linha, status="Feito", intervalo=10):
    """Marca o status e salva o arquivo a cada N atualizações."""
    ws = wb.active
    ws[f'F{linha}'] = status
    contador += 1
    if contador % intervalo == 0:
        wb.save(caminho_planilha)
        print(f"[PLANILHA] Salvamento parcial após {contador} atualizações (última linha {linha}).")
    return contador


def main():
    """Função principal"""
    cpfs_para_processar = ler_cpfs_planilha(PLANILHA_PATH)
    total_cpfs = len(cpfs_para_processar)
    
    if not cpfs_para_processar:
        print("\n[ALERTA] Nenhum CPF pendente na planilha.")
        return
    
    print("="*60)
    print("VALIDADOR DE COLABORADORES")
    print("="*60)
    print(f"\n[INFO] Total de CPFs a processar: {total_cpfs}")
    print("\nIniciando navegador...")
    
    # Configura navegador
    options = uc.ChromeOptions()
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--disable-save-password-bubble")
    options.add_argument("--disable-features=PasswordManager,PasswordManagerOnboarding")
    
    driver = uc.Chrome(options=options, headless=False)
    wb_atualizacao = None
    atualizacoes_desde_save = 0
    
    try:
        wb_atualizacao = load_workbook(PLANILHA_PATH)
    except Exception as e:
        print(f"[ERRO] Não foi possível abrir a planilha: {e}")
        print("[ALERTA] Feche o arquivo no Excel para permitir atualização.")
        wb_atualizacao = None
    
    try:
        print("\n[NAVEGAÇÃO] Abrindo página de login...")
        driver.get(LOGIN_URL)
        time.sleep(3)
        
        print("\n" + "="*60)
        print("ATENÇÃO: Faça login e navegue manualmente até")
        print("'Colaborador Avançado' e pressione Enter...")
        print("="*60)
        input("\nPressione Enter quando estiver pronto...")
        
        print("\n[VALIDAÇÃO] Iniciando processamento...")
        
        sucessos = 0
        falhas = 0
        
        for i, item in enumerate(cpfs_para_processar, 1):
            cpf = item['cpf']
            linha = item['linha']
            
            print(f"\n{'='*60}")
            print(f"[VALIDAÇÃO] Processando {i}/{total_cpfs}")
            print(f"[VALIDAÇÃO] Linha {linha} - CPF {cpf}")
            print(f"{'='*60}")
            
            if validar_colaborador(driver, cpf):
                sucessos += 1
                if wb_atualizacao:
                    atualizacoes_desde_save = atualizar_status_planilha_batch(
                        wb_atualizacao, PLANILHA_PATH, atualizacoes_desde_save, linha, "Feito", 10
                    )
            else:
                falhas += 1
                if wb_atualizacao:
                    atualizacoes_desde_save = atualizar_status_planilha_batch(
                        wb_atualizacao, PLANILHA_PATH, atualizacoes_desde_save, linha, "Erro", 10
                    )
            
            time.sleep(2)
        
        # Resumo
        print(f"\n{'='*60}")
        print(f"[RESUMO] ✓ Sucessos: {sucessos} | ✗ Falhas: {falhas}")
        print(f"{'='*60}\n")
        
    except Exception as e:
        print(f"\n[ERRO] {e}")
    finally:
        if wb_atualizacao:
            try:
                wb_atualizacao.save(PLANILHA_PATH)
                wb_atualizacao.close()
                print("[PLANILHA] Salvo.")
            except Exception as e:
                print(f"[ERRO] Falha ao salvar: {e}")
        print("\nPressione Enter para fechar...")
        input()
        driver.quit()


if __name__ == "__main__":
    main()
